#!/usr/bin/env python3
"""Helpers for personal balance-sheet workbooks."""

from __future__ import annotations

import argparse
import posixpath
import re
import shutil
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:  # pragma: no cover - operator guidance path
    raise SystemExit(
        "openpyxl is required. Create a temporary venv and install openpyxl, "
        "then rerun this script."
    ) from exc


CATEGORY_SHEETS = ["流动资产", "投资资产", "应收账款", "负债"]
SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".xls", ".csv"}
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
TEXT_UNIT_REPLACEMENTS = [
    ("单位/W", "单位/元"),
    ("单位：万元", "单位：元"),
    ("单位：万", "单位：元"),
    ("/W", "/元"),
    ("(W)", "(元)"),
    ("（W）", "（元）"),
    ("万元；", "元；"),
    ("万元", "元"),
]
NOTE_UNIT_REPLACEMENTS = [
    ("已还金额0.75", "已还金额7500元"),
    ("2.5+0.2", "25000+2000元"),
    ("已还金额0.2，剩余0.3W", "已还金额2000元，剩余3000元"),
    ("剩余0.3W", "剩余3000元"),
]


def normalize_path(raw: str) -> Path:
    text = raw.strip().strip('"')
    prefix = "\\\\wsl.localhost\\Ubuntu"
    if text.startswith(prefix):
        text = text[len(prefix) :].replace("\\", "/")
        return Path(text)
    if len(text) >= 3 and text[1:3] == ":\\":
        drive = text[0].lower()
        return Path(f"/mnt/{drive}/{text[3:].replace('\\', '/')}")
    return Path(text)


def workbook_zip_ok(path: Path) -> bool:
    with ZipFile(path) as archive:
        return archive.testzip() is None


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def spreadsheet_files(directory: Path):
    return sorted(
        path
        for path in directory.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and any(path.name.endswith(suffix) for suffix in SPREADSHEET_SUFFIXES)
    )


def detail_image_files(directory: Path, recursive: bool = False):
    iterator = directory.rglob("*") if recursive else directory.iterdir()
    return sorted(
        path
        for path in iterator
        if path.is_file()
        and not path.name.startswith(".")
        and path.suffix.lower() in IMAGE_SUFFIXES
    )


def destination_folder_name(filename: str) -> str | None:
    month_match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", filename)
    if month_match:
        return f"{month_match.group(1)}.{int(month_match.group(2))}"
    year_match = re.search(r"(20\d{2})", filename)
    if year_match:
        return year_match.group(1)
    return None


def move_with_zone_identifier(source: Path, destination_dir: Path) -> list[tuple[Path, Path]]:
    moved = []
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / source.name
    if destination.exists():
        raise SystemExit(f"Refusing to overwrite existing file: {destination}")
    shutil.move(str(source), str(destination))
    moved.append((source, destination))

    zone_source = Path(f"{source}:Zone.Identifier")
    if zone_source.exists():
        zone_destination = destination_dir / zone_source.name
        if zone_destination.exists():
            raise SystemExit(f"Refusing to overwrite existing file: {zone_destination}")
        shutil.move(str(zone_source), str(zone_destination))
        moved.append((zone_source, zone_destination))
    return moved


def ocr_image_text(path: Path) -> str:
    try:
        result = subprocess.run(
            ["tesseract", str(path), "stdout", "-l", "chi_sim+eng", "--psm", "6"],
            check=False,
            capture_output=True,
            text=True,
            timeout=30,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return path.name
    return f"{path.name}\n{result.stdout}"


def classify_detail_image(path: Path, text: str) -> str | None:
    haystack = f"{path.name}\n{text}"
    filename_rules = [
        ("有钱花", "有钱花"),
        ("借呗", "借呗"),
        ("招行贷款", "招行"),
        ("民生贷款", "民生"),
        ("浦发贷款", "浦发"),
    ]
    for folder, keyword in filename_rules:
        if keyword in path.stem:
            return folder
    rules = [
        ("云闪付", ("余额查询", "可用余额总计", "可查询储")),
        ("有钱花", ("有钱花", "查账", "共4笔借款", "06月08日应还", "144, 500")),
        ("借呗", ("借呗", "先息后本", "剩余的17000", "剩余的150000")),
        ("招行贷款", ("招行", "剩余待还本金", "06-18")),
        ("民生贷款", ("民易贷", "我要还款")),
        ("浦发贷款", ("浦发", "2027/01/29", "90,244.13")),
        ("证券账户", ("东方证券", "东方财富证券", "长江证券")),
        ("天天基金", ("天天基金", "活期宝", "投顾", "银行黄金")),
        ("微众银行理财", ("微众", "我的资产(元)", "进阶资产", "会员中心")),
        ("支付宝", ("支付宝", "余额宝", "稳健理财", "指数+")),
    ]
    for folder, keywords in rules:
        if any(keyword in haystack for keyword in keywords):
            return folder
    return None


def safe_rename(source: Path, destination: Path) -> Path:
    if source == destination:
        return destination
    destination.parent.mkdir(parents=True, exist_ok=True)
    if not destination.exists():
        source.rename(destination)
        return destination
    stem = destination.stem
    suffix = destination.suffix
    for index in range(2, 100):
        candidate = destination.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            source.rename(candidate)
            return candidate
    raise SystemExit(f"Cannot find available destination for: {destination}")


def amount_token(value: str) -> str:
    return value.replace(",", "").replace(" ", "")


def first_amount_after(text: str, label: str) -> str | None:
    index = text.find(label)
    if index < 0:
        return None
    match = re.search(r"([0-9][0-9,\s]*\.\d{2}|[0-9][0-9,\s]*)", text[index + len(label) : index + len(label) + 80])
    return amount_token(match.group(1)) if match else None


def amounts_in_text(text: str) -> list[str]:
    return [amount_token(match) for match in re.findall(r"[0-9][0-9,\s]*\.\d{2}", text)]


def readable_detail_name(path: Path, text: str) -> str | None:
    folder = path.parent.name
    amounts = amounts_in_text(text)
    suffix = path.suffix
    if folder == "云闪付":
        amount = first_amount_after(text, "可用余额总计") or (amounts[0] if amounts else None)
        return f"云闪付_可用余额总计{amount}元{suffix}" if amount else None
    if folder == "借呗":
        principal = first_amount_after(text, "剩余的") or (amounts[0] if amounts else None)
        due = amounts[-1] if amounts else None
        if principal and due and due != principal:
            return f"借呗_本金{principal}元_应还{due}元{suffix}"
        return f"借呗_本金{principal}元{suffix}" if principal else None
    if folder == "天天基金":
        if len(amounts) >= 2:
            total = amount_token(str(round(float(amounts[0]) + float(amounts[1]), 2)))
            return f"天天基金_总资产{total}元_活期宝{amounts[0]}_基金{amounts[1]}{suffix}"
    if folder == "微众银行理财":
        total = first_amount_after(text, "我的资产(元)") or (amounts[0] if amounts else None)
        if total and len(amounts) >= 3:
            return f"微众银行理财_总资产{total}元_活期{amounts[1]}_进阶{amounts[2]}{suffix}"
        return f"微众银行理财_总资产{total}元{suffix}" if total else None
    if folder == "招行贷款":
        principal = first_amount_after(text, "剩余待还本金") or (amounts[0] if amounts else None)
        interest = None
        if amounts:
            interest_values = [float(value) for value in amounts if float(value) < 1000]
            if interest_values:
                interest = amount_token(str(round(sum(interest_values), 2)))
        if principal and interest:
            return f"招行贷款_剩余本金{principal}元_利息{interest}元{suffix}"
        return f"招行贷款_剩余本金{principal}元{suffix}" if principal else None
    if folder == "支付宝":
        amount = first_amount_after(text, "总资产") or (amounts[0] if amounts else None)
        return f"支付宝理财_总资产{amount}元{suffix}" if amount else None
    if folder == "有钱花":
        due = first_amount_after(text, "应还") or (amounts[0] if amounts else None)
        principal = first_amount_after(text, "待还本金") or first_amount_after(text, "本金")
        interest = first_amount_after(text, "利息")
        if principal and interest and due:
            return f"有钱花_应还{due}元_本金{principal}_利息{interest}{suffix}"
        if due and principal:
            return f"有钱花_应还{due}元_待还本金{principal}元{suffix}"
        return f"有钱花_应还{due}元{suffix}" if due else None
    if folder == "民生贷款":
        principal = first_amount_after(text, "剩余本金") or (amounts[1] if len(amounts) > 1 else None)
        due = first_amount_after(text, "06-16") or (amounts[2] if len(amounts) > 2 else None)
        if principal and due:
            return f"民生贷款_剩余本金{principal}元_06月应还{due}元{suffix}"
        return f"民生贷款_剩余本金{principal}元{suffix}" if principal else None
    if folder == "浦发贷款":
        principal = first_amount_after(text, "本金") or "90000"
        due = first_amount_after(text, "2027/01/29") or (amounts[-1] if amounts else None)
        if principal and due:
            return f"浦发贷款_剩余本金{principal}元_到期应还{due}元{suffix}"
    if folder == "证券账户":
        if len(amounts) >= 3:
            return f"证券账户_东方{amounts[0]}元_东财{amounts[1]}_长江{amounts[2]}{suffix}"
    return None


def replace_unit_texts(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    text = cell.value
                    for old, new in TEXT_UNIT_REPLACEMENTS + NOTE_UNIT_REPLACEMENTS:
                        text = text.replace(old, new)
                    cell.value = text
                if cell.comment and cell.comment.text:
                    text = cell.comment.text
                    for old, new in TEXT_UNIT_REPLACEMENTS + NOTE_UNIT_REPLACEMENTS:
                        text = text.replace(old, new)
                    cell.comment.text = text


def sheet_has_w_unit(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and (
                "/W" in value or "(W)" in value or "（W）" in value or "万元" in value or "单位/W" in value
            ):
                return True
    return False


def multiply_cell_by_10000(ws, coord: str, changed: list[str]) -> None:
    cell = ws[coord]
    if is_number(cell.value):
        cell.value = round(cell.value * 10000, 6)
        cell.number_format = "#,##0.00"
        changed.append(f"{ws.title}!{coord}")


def multiply_range_by_10000(ws, rows, columns: list[str], changed: list[str]) -> None:
    for row in rows:
        for column in columns:
            multiply_cell_by_10000(ws, f"{column}{row}", changed)


def set_recalculate_on_open(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcMode = "auto"


def money_value(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if is_number(value):
        return float(value)
    raise SystemExit(f"Expected numeric amount, got: {value!r}")


def rounded_money(value: float) -> float:
    return round(value + 0.0000001, 2)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def unmerge_from_row(ws, start_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= start_row:
            ws.unmerge_cells(str(merged))


def normalize_workbook_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def patch_formula_caches(workbook: Path, sheet_caches: dict[str, dict[str, float]]) -> None:
    """Write cached numeric values for formula cells so data_only readers see totals."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ET.register_namespace("", main_ns)
    ET.register_namespace("r", rel_ns)

    with ZipFile(workbook, "r") as source_archive:
        workbook_root = ET.fromstring(source_archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(source_archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {}
        for relation in rels_root:
            relation_id = relation.attrib.get("Id")
            target = relation.attrib.get("Target")
            if relation_id and target:
                rel_targets[relation_id] = normalize_workbook_target(target)

        sheet_paths = {}
        sheets = workbook_root.find(f"{{{main_ns}}}sheets")
        if sheets is not None:
            for sheet in sheets:
                name = sheet.attrib.get("name")
                relation_id = sheet.attrib.get(f"{{{rel_ns}}}id")
                target = rel_targets.get(relation_id or "")
                if name and target:
                    sheet_paths[name] = target

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = Path(temp_file.name)

        with ZipFile(temp_path, "w", compression=ZIP_DEFLATED) as target_archive:
            for item in source_archive.infolist():
                data = source_archive.read(item.filename)
                sheet_name = next(
                    (name for name, sheet_path in sheet_paths.items() if sheet_path == item.filename),
                    None,
                )
                if sheet_name in sheet_caches:
                    root = ET.fromstring(data)
                    cells_by_ref = {
                        cell.attrib.get("r"): cell
                        for cell in root.iter(f"{{{main_ns}}}c")
                        if cell.attrib.get("r")
                    }
                    for ref, value in sheet_caches[sheet_name].items():
                        cell = cells_by_ref.get(ref)
                        if cell is None or cell.find(f"{{{main_ns}}}f") is None:
                            continue
                        cached = cell.find(f"{{{main_ns}}}v")
                        if cached is None:
                            cached = ET.SubElement(cell, f"{{{main_ns}}}v")
                        cached.text = str(rounded_money(value))
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                target_archive.writestr(item, data)

    shutil.move(str(temp_path), workbook)


def first_main_sheet(wb, requested: str | None = None):
    if requested:
        return wb[requested]
    for preferred in ("202603", "个人资产负债表"):
        if preferred in wb.sheetnames:
            return wb[preferred]
    return wb[wb.sheetnames[0]]


def copy_template_sheet(source_wb, target_wb, template_name: str):
    if template_name not in source_wb.sheetnames:
        return target_wb.create_sheet(template_name)

    template = source_wb[template_name]
    ws = target_wb.create_sheet(template_name)
    for row in template.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            new_cell = ws[cell.coordinate]
            if cell.has_style:
                new_cell._style = copy(cell._style)
            new_cell.number_format = cell.number_format
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)

    for key, dim in template.column_dimensions.items():
        ws.column_dimensions[key].width = dim.width
        ws.column_dimensions[key].hidden = dim.hidden
    for key, dim in template.row_dimensions.items():
        ws.row_dimensions[key].height = dim.height
        ws.row_dimensions[key].hidden = dim.hidden
    return ws


def clear_values(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.comment = None


def style_generated_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None:
                continue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for row_idx in (1, 2, 3, 16):
        if row_idx > ws.max_row:
            continue
        for cell in ws[row_idx]:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            cell.font = Font(bold=True, size=12 if row_idx == 1 else 11)
            cell.fill = section_fill if row_idx == 1 else header_fill


def set_number_format(ws, cells: list[str], number_format: str = "0.00") -> None:
    for coord in cells:
        ws[coord].number_format = number_format


def split_categories(workbook: Path, reference: Path | None, main_sheet: str | None) -> None:
    wb = load_workbook(workbook, data_only=False)
    main = first_main_sheet(wb, main_sheet)
    source_wb = load_workbook(reference, data_only=False) if reference else None

    for sheet_name in CATEGORY_SHEETS:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    def make_sheet(name: str):
        if source_wb:
            ws = copy_template_sheet(source_wb, wb, name)
        else:
            ws = wb.create_sheet(name)
        clear_values(ws)
        return ws

    ws = make_sheet("流动资产")
    ws["B1"] = "流动资产汇总"
    ws["A2"] = "账户类型"
    ws["B2"] = "账户"
    ws["C2"] = "活期"
    ws["D2"] = "基金/理财"
    ws["E2"] = "公积金"
    liquid_rows = [
        ("银行卡", main["B2"].value, main["C2"].value, None, None),
        ("银行卡", main["B3"].value, main["C3"].value, None, None),
        ("公积金", main["B4"].value, None, None, main["C4"].value),
    ]
    for row_idx, row_values in enumerate(liquid_rows, start=3):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row_idx, col_idx).value = value
            if col_idx >= 3:
                ws.cell(row_idx, col_idx).number_format = "0.00"
    ws["B11"] = "累计(W)"
    ws["C11"] = "=SUM(C3:C10)"
    ws["D11"] = "=SUM(D3:D10)"
    ws["E11"] = "=SUM(E3:E10)"
    ws["B12"] = "流动资产合计(W)"
    ws["C12"] = "=SUM(C11:E11)"
    set_number_format(ws, ["C11", "D11", "E11", "C12"])

    ws = make_sheet("投资资产")
    for target_row, source_row in enumerate(range(5, 13), start=1):
        ws.cell(target_row, 1).value = "投资账户" if target_row == 1 else None
        ws.cell(target_row, 2).value = main.cell(source_row, 2).value
        ws.cell(target_row, 3).value = main.cell(source_row, 3).value
        ws.cell(target_row, 3).number_format = "0.00"
    ws["B10"] = "累计(W)"
    ws["C10"] = "=SUM(C1:C9)"
    ws["C10"].number_format = "0.00"

    ws = make_sheet("应收账款")
    ws["A1"] = "应收账款明细(W)"
    headers = ["借款人", "借款金额/W", "借款时间", "备注"]
    for start_col in (1, 6, 11):
        for offset, header in enumerate(headers):
            ws.cell(2, start_col + offset).value = header
    for start_col in (1, 6):
        for offset, header in enumerate(headers):
            ws.cell(16, start_col + offset).value = header
    for start_col, target_row, source_row in (
        (1, 3, 14),
        (6, 3, 15),
        (11, 3, 16),
        (1, 17, 17),
        (6, 17, 18),
    ):
        ws.cell(target_row, start_col).value = main.cell(source_row, 2).value
        ws.cell(target_row, start_col + 1).value = main.cell(source_row, 3).value
        ws.cell(target_row, start_col + 1).number_format = "0.00"
        ws.cell(target_row, start_col + 2).value = main.cell(source_row, 4).value
        ws.cell(target_row, start_col + 3).value = main.cell(source_row, 9).value
    ws["A14"] = "剩余欠款(W)"
    ws["B14"] = "=SUM(B3:B13)"
    ws["F14"] = "剩余欠款(W)"
    ws["G14"] = "=SUM(G3:G13)"
    ws["K14"] = "剩余欠款(W)"
    ws["L14"] = "=SUM(L3:L13)"
    ws["A20"] = "剩余欠款(W)"
    ws["B20"] = "=SUM(B17:B19)"
    ws["F20"] = "剩余欠款(W)"
    ws["G20"] = "=SUM(G17:G19)"
    ws["K16"] = "应收账款(W)"
    ws["L16"] = "=SUM(B14,G14,L14,B20,G20)"
    set_number_format(ws, ["B14", "G14", "L14", "B20", "G20", "L16"])

    ws = make_sheet("负债")
    ws["A1"] = "银行网贷借贷详情(W)"
    headers = ["项目", "借款金额/W", "年利率", "待归还本金/W", "待归还利息/W", "累计利息/元", "截至日期", "备注"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(3, col_idx).value = header
    for target_row, source_row in enumerate(range(20, 25), start=4):
        note = main.cell(source_row, 6).comment.text if main.cell(source_row, 6).comment else None
        values = [
            main.cell(source_row, 2).value,
            main.cell(source_row, 3).value,
            main.cell(source_row, 4).value,
            main.cell(source_row, 5).value,
            main.cell(source_row, 6).value,
            main.cell(source_row, 8).value,
            main.cell(source_row, 9).value,
            note,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(target_row, col_idx).value = value
        for col_idx in (2, 4, 5):
            ws.cell(target_row, col_idx).number_format = "0.00"
        ws.cell(target_row, 3).number_format = "0.00%"
        ws.cell(target_row, 6).number_format = "0.00"
        ws.cell(target_row, 7).number_format = "yyyy/m/d"
    ws["A10"] = "合计"
    ws["D10"] = "=SUM(D4:D8)"
    ws["E10"] = "=SUM(E4:E8)"
    ws["F10"] = "=SUM(F4:F8)"
    ws["H10"] = "注：截图未展示全部剩余利息的负债，保留空值。"
    set_number_format(ws, ["D10", "E10", "F10"])

    for sheet_name in CATEGORY_SHEETS:
        style_generated_sheet(wb[sheet_name])

    ordered = [main] + [wb[name] for name in CATEGORY_SHEETS]
    wb._sheets = ordered + [sheet for sheet in wb.worksheets if sheet not in ordered]
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcMode = "auto"
    wb.save(workbook)


def archive_by_date(directory: Path) -> list[tuple[Path, Path]]:
    moved: list[tuple[Path, Path]] = []
    for source in spreadsheet_files(directory):
        folder = destination_folder_name(source.name)
        if not folder:
            print(f"skip_no_date: {source.name}")
            continue
        moved.extend(move_with_zone_identifier(source, directory / folder))
    return moved


def receivable_records_from_main(main):
    records = []
    for row in range(14, 19):
        borrower = main.cell(row, 2).value
        if not borrower:
            continue
        records.append(
            {
                "borrower": str(borrower),
                "amount": main.cell(row, 3).value,
                "date": main.cell(row, 4).value,
                "repaid": main.cell(row, 5).value,
                "remaining": main.cell(row, 6).value,
                "note": main.cell(row, 9).value,
                "source_row": row,
            }
        )
    return records


def receivable_records_from_summary(wb):
    if "汇总" not in wb.sheetnames:
        return []
    summary = wb["汇总"]
    records = []
    for row in range(3, summary.max_row + 1):
        borrower = summary.cell(row, 1).value
        if not borrower or borrower == "合计":
            continue
        records.append(
            {
                "borrower": str(borrower),
                "amount": summary.cell(row, 2).value,
                "date": summary.cell(row, 3).value,
                "repaid": summary.cell(row, 4).value,
                "remaining": summary.cell(row, 5).value,
                "note": summary.cell(row, 6).value,
                "source": summary.cell(row, 7).value,
            }
        )
    return records


def to_actual_amount(value, source_is_w: bool):
    if is_number(value):
        scale = 10000 if source_is_w else 1
        return round(value * scale, 2)
    return None


def style_table(ws, max_row: int, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = Font(bold=True, size=14)
            cell.fill = title_fill
    for cell in ws[2]:
        if cell.value is not None:
            cell.font = Font(bold=True)
            cell.fill = header_fill


def month_prefix_from_path(path: Path) -> str | None:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", path.name)
    return match.group(0) if match else None


def month_title(prefix: str | None, suffix: str) -> str:
    if not prefix:
        return suffix
    return f"{prefix[:4]}年{int(prefix[4:])}月{suffix}"


def infer_month_workbook(directory: Path, keyword: str) -> Path:
    candidates = [
        path
        for path in sorted(directory.glob("*.xlsx"))
        if not path.name.startswith("~$") and keyword in path.name
    ]
    if not candidates:
        raise SystemExit(f"No *{keyword}*.xlsx workbook found in {directory}")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise SystemExit(f"Multiple *{keyword}*.xlsx workbooks found, pass an explicit path: {names}")
    return candidates[0]


def infer_asset_workbook(directory: Path) -> Path:
    candidates = [
        path
        for path in sorted(directory.glob("*.xlsx"))
        if not path.name.startswith("~$")
        and ("资产表" in path.name or "资产负债表" in path.name)
        and "应收账款" not in path.name
        and "负债表" not in path.name
    ]
    if not candidates:
        raise SystemExit(f"No asset workbook found in {directory}")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise SystemExit(f"Multiple asset workbooks found, pass --workbook: {names}")
    return candidates[0]


def read_receivable_summary(workbook: Path) -> list[dict]:
    wb = load_workbook(workbook, data_only=True)
    if "汇总" not in wb.sheetnames:
        raise SystemExit(f"Receivable workbook has no 汇总 sheet: {workbook}")
    ws = wb["汇总"]
    records = []
    for row in range(3, ws.max_row + 1):
        borrower = ws.cell(row, 1).value
        if not borrower or borrower == "合计":
            continue
        amount = money_value(ws.cell(row, 2).value)
        repaid = money_value(ws.cell(row, 4).value)
        cached_remaining = ws.cell(row, 5).value
        remaining = money_value(cached_remaining) if is_number(cached_remaining) and cached_remaining else amount - repaid
        records.append(
            {
                "borrower": str(borrower),
                "amount": rounded_money(amount),
                "date": ws.cell(row, 3).value,
                "repaid": rounded_money(repaid),
                "remaining": rounded_money(remaining),
                "note": ws.cell(row, 6).value,
                "source": ws.cell(row, 7).value,
            }
        )
    if not records:
        raise SystemExit(f"No receivable records found in {workbook}")
    return records


def read_liability_summary(workbook: Path) -> list[dict]:
    wb = load_workbook(workbook, data_only=True)
    if "汇总" not in wb.sheetnames:
        raise SystemExit(f"Liability workbook has no 汇总 sheet: {workbook}")
    ws = wb["汇总"]
    records = []
    for row in range(3, ws.max_row + 1):
        platform = ws.cell(row, 1).value
        if not platform or platform == "合计":
            continue
        principal = money_value(ws.cell(row, 4).value)
        interest = money_value(ws.cell(row, 5).value)
        known_total = ws.cell(row, 6).value
        records.append(
            {
                "platform": str(platform),
                "borrowed": rounded_money(money_value(ws.cell(row, 2).value)),
                "rate": ws.cell(row, 3).value,
                "principal": rounded_money(principal),
                "interest": rounded_money(interest),
                "known_total": rounded_money(money_value(known_total) if is_number(known_total) else principal + interest),
                "accrued_interest": rounded_money(money_value(ws.cell(row, 7).value)),
                "due_at": ws.cell(row, 8).value,
                "note": ws.cell(row, 9).value,
                "source": ws.cell(row, 10).value,
            }
        )
    if not records:
        raise SystemExit(f"No liability records found in {workbook}")
    return records


def sum_numeric(ws, cells: list[str]) -> float:
    return sum(money_value(ws[coord].value) for coord in cells if is_number(ws[coord].value))


def update_main_from_summaries(
    workbook: Path,
    receivables_workbook: Path,
    liabilities_workbook: Path,
    main_sheet: str | None,
) -> None:
    receivables = read_receivable_summary(receivables_workbook)
    liabilities = read_liability_summary(liabilities_workbook)
    month_prefix = (
        month_prefix_from_path(workbook)
        or month_prefix_from_path(receivables_workbook)
        or month_prefix_from_path(liabilities_workbook)
    )
    receivable_total = rounded_money(sum(record["remaining"] for record in receivables))
    liability_total = rounded_money(sum(record["known_total"] for record in liabilities))

    wb = load_workbook(workbook, data_only=False)
    main = first_main_sheet(wb, main_sheet)
    if main.max_row < 12:
        raise SystemExit(f"Main sheet has too few rows to update safely: {main.title}")

    header_styles = {col: copy(main.cell(2, col)._style) for col in range(1, 11)}
    item_styles = {col: copy(main.cell(6, col)._style) for col in range(1, 11)}
    total_styles = {col: copy(main.cell(main.max_row, col)._style) for col in range(1, 11)}
    category_cell = main["A5"]

    unmerge_from_row(main, 13)
    if main.max_row >= 13:
        main.delete_rows(13, main.max_row - 12)

    rows = [["应收账款", "借款人", "表内应收金额/元", "借款时间", "已还金额/元", "待还金额/元", None, None, "备注", receivable_total]]
    rows.extend(
        [
            None,
            record["borrower"],
            record["amount"],
            record["date"],
            record["repaid"],
            record["remaining"],
            None,
            None,
            record["note"],
            None,
        ]
        for record in receivables
    )
    rows.append(
        [
            "负债",
            "负债平台",
            "借款金额/元",
            "年利率",
            "待归还本金/元",
            "待归还利息/元",
            "已知待还合计/元",
            "累计利息/元",
            "截至日期",
            -liability_total,
        ]
    )
    rows.extend(
        [
            None,
            record["platform"],
            record["borrowed"],
            record["rate"],
            record["principal"],
            record["interest"],
            record["known_total"],
            record["accrued_interest"],
            record["due_at"],
            None,
        ]
        for record in liabilities
    )

    start_row = 13
    net_row = start_row + len(rows)
    rows.append([f"=SUM(J2:J{net_row - 1})", None, None, None, None, None, None, None, None, None])
    for row_offset, values in enumerate(rows):
        row = start_row + row_offset
        is_header = row_offset in (0, len(receivables) + 1)
        is_total = row_offset == len(rows) - 1
        for col, value in enumerate(values, start=1):
            cell = main.cell(row, col)
            cell.value = value
            if is_total:
                cell._style = copy(total_styles[col])
            elif is_header:
                cell._style = copy(header_styles[col])
            else:
                cell._style = copy(item_styles[col])

    receivable_header = start_row
    receivable_end = receivable_header + len(receivables)
    liability_header = receivable_end + 1
    liability_end = liability_header + len(liabilities)
    main.merge_cells(start_row=receivable_header, start_column=1, end_row=receivable_end, end_column=1)
    main.merge_cells(start_row=receivable_header, start_column=10, end_row=receivable_end, end_column=10)
    main.merge_cells(start_row=liability_header, start_column=1, end_row=liability_end, end_column=1)
    main.merge_cells(start_row=liability_header, start_column=10, end_row=liability_end, end_column=10)
    main.merge_cells(start_row=net_row, start_column=1, end_row=net_row, end_column=10)
    copy_cell_style(category_cell, main.cell(receivable_header, 1))
    copy_cell_style(category_cell, main.cell(liability_header, 1))

    for row in range(receivable_header, receivable_end + 1):
        for col in (3, 5, 6, 10):
            main.cell(row, col).number_format = "#,##0.00"
        main.cell(row, 4).number_format = "yyyy/m/d"
    for row in range(liability_header, liability_end + 1):
        for col in (3, 5, 6, 7, 8, 10):
            main.cell(row, col).number_format = "#,##0.00"
        main.cell(row, 4).number_format = "0.00%"
        main.cell(row, 9).number_format = "yyyy/m/d"
    main.cell(net_row, 1).number_format = "#,##0.00"
    for column, width in {"A": 13, "B": 16, "C": 16, "D": 13, "E": 16, "F": 16, "G": 18, "H": 16, "I": 28, "J": 18}.items():
        main.column_dimensions[column].width = width

    for sheet_name in ("应收账款", "负债"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    receivable_sheet = wb.create_sheet("应收账款")
    receivable_sheet.append([month_title(month_prefix, "应收账款汇总")])
    receivable_sheet.append(["借款人", "表内应收金额/元", "借款时间", "已还金额/元", "待还金额/元", "备注", "来源"])
    for record in receivables:
        receivable_sheet.append(
            [record["borrower"], record["amount"], record["date"], record["repaid"], record["remaining"], record["note"], record["source"]]
        )
    receivable_sheet.append(
        ["合计", rounded_money(sum(record["amount"] for record in receivables)), None, rounded_money(sum(record["repaid"] for record in receivables)), receivable_total, None, None]
    )
    receivable_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    for column, width in {"A": 14, "B": 18, "C": 14, "D": 16, "E": 16, "F": 34, "G": 28}.items():
        receivable_sheet.column_dimensions[column].width = width
    for row in range(3, receivable_sheet.max_row + 1):
        for col in (2, 4, 5):
            receivable_sheet.cell(row, col).number_format = "#,##0.00"
        receivable_sheet.cell(row, 3).number_format = "yyyy/m/d"
    style_table(receivable_sheet, receivable_sheet.max_row, 7)

    liability_sheet = wb.create_sheet("负债")
    liability_sheet.append([month_title(month_prefix, "负债汇总")])
    liability_sheet.append(["负债平台", "借款金额/元", "年利率", "待归还本金/元", "待归还利息/元", "已知待还合计/元", "累计利息/元", "截至日期", "备注", "来源"])
    for record in liabilities:
        liability_sheet.append(
            [
                record["platform"],
                record["borrowed"],
                record["rate"],
                record["principal"],
                record["interest"],
                record["known_total"],
                record["accrued_interest"],
                record["due_at"],
                record["note"],
                record["source"],
            ]
        )
    liability_sheet.append(
        [
            "合计",
            rounded_money(sum(record["borrowed"] for record in liabilities)),
            None,
            rounded_money(sum(record["principal"] for record in liabilities)),
            rounded_money(sum(record["interest"] for record in liabilities)),
            liability_total,
            rounded_money(sum(record["accrued_interest"] for record in liabilities)),
            None,
            None,
            None,
        ]
    )
    liability_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    for column, width in {"A": 14, "B": 16, "C": 12, "D": 18, "E": 18, "F": 18, "G": 16, "H": 14, "I": 42, "J": 32}.items():
        liability_sheet.column_dimensions[column].width = width
    for row in range(3, liability_sheet.max_row + 1):
        for col in (2, 4, 5, 6, 7):
            liability_sheet.cell(row, col).number_format = "#,##0.00"
        liability_sheet.cell(row, 3).number_format = "0.00%"
        liability_sheet.cell(row, 8).number_format = "yyyy/m/d"
    style_table(liability_sheet, liability_sheet.max_row, 10)

    liquid_total = sum_numeric(main, ["C2", "C3", "C4"])
    investment_total = sum(
        money_value(main.cell(row, 3).value)
        for row in range(5, 13)
        if is_number(main.cell(row, 3).value)
    )
    net_assets = rounded_money(liquid_total + investment_total + receivable_total - liability_total)
    set_recalculate_on_open(wb)
    wb.save(workbook)

    patch_formula_caches(
        workbook,
        {
            main.title: {"J2": liquid_total, "J5": investment_total, f"A{net_row}": net_assets},
            "流动资产": {"C11": sum_numeric(wb["流动资产"], ["C3", "C4"]), "D11": 0, "E11": money_value(wb["流动资产"]["E5"].value), "C12": liquid_total}
            if "流动资产" in wb.sheetnames
            else {},
            "投资资产": {"C10": investment_total} if "投资资产" in wb.sheetnames else {},
        },
    )

    print(f"updated: {workbook}")
    print(f"receivable_total: {receivable_total:.2f}")
    print(f"liability_total: {liability_total:.2f}")
    print(f"net_assets: {net_assets:.2f}")


def write_receivable_ledger_sheet(ws, record: dict, source_name: str | None, source_row: int | None, source_is_w: bool) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{record['borrower']}应收账款"
    ws["A2"] = "时间"
    ws["B2"] = "借/还"
    ws["C2"] = "金额"
    ws["D2"] = "剩余待还"

    amount_yuan = to_actual_amount(record.get("amount"), source_is_w)
    repaid_yuan = to_actual_amount(record.get("repaid"), source_is_w)
    ws["A3"] = record.get("date")
    ws["B3"] = "借"
    ws["C3"] = amount_yuan
    ws["D3"] = amount_yuan

    next_row = 4
    if repaid_yuan:
        ws.cell(next_row, 2).value = "还"
        ws.cell(next_row, 3).value = repaid_yuan
        ws.cell(next_row, 4).value = f"=D{next_row - 1}-C{next_row}"
        next_row += 1

    note_parts = []
    if record.get("note"):
        note_parts.append(str(record["note"]))
    if source_name and source_row:
        note_parts.append(f"来源：{source_name}!{source_row}行")
    elif record.get("source"):
        note_parts.append(f"来源：{record['source']}")
    if note_parts:
        ws.cell(next_row + 1, 1).value = "备注"
        ws.cell(next_row + 1, 2).value = "；".join(note_parts)

    for column, width in {"A": 16, "B": 12, "C": 16, "D": 16}.items():
        ws.column_dimensions[column].width = width
    for row in range(3, next_row + 2):
        ws.cell(row, 1).number_format = "yyyy/m/d"
        ws.cell(row, 3).number_format = "#,##0.00"
        ws.cell(row, 4).number_format = "#,##0.00"
    style_table(ws, max(next_row + 1, 4), 4)


def export_receivables(workbook: Path, output: Path | None, main_sheet: str | None, ledger: bool) -> Path:
    source_wb = load_workbook(workbook, data_only=False)
    main = first_main_sheet(source_wb, main_sheet)
    source_is_w = sheet_has_w_unit(main)
    records = receivable_records_from_main(main)
    if not records:
        raise SystemExit("No receivable rows found in the main sheet (expected rows 14-18).")
    output = output or workbook.with_name(f"{workbook.stem.replace('资产负债表', '')}应收账款.xlsx")

    wb = Workbook()
    summary = wb.active
    summary.title = "汇总"
    summary["A1"] = "应收账款汇总"
    unit = "W" if source_is_w else "元"
    summary.append(["借款人", f"表内应收金额/{unit}", "借款时间", f"已还金额/{unit}", f"待还金额/{unit}", "备注", "来源"])
    for record in records:
        summary.append(
            [
                record["borrower"],
                record["amount"],
                record["date"],
                record["repaid"],
                record["remaining"],
                record["note"],
                f"{workbook.name}!{record['source_row']}行",
            ]
        )
    end_row = len(records) + 2
    summary.append(["合计", f"=SUM(B3:B{end_row})", None, f"=SUM(D3:D{end_row})", f"=SUM(E3:E{end_row})", None, None])
    for column, width in {"A": 14, "B": 18, "C": 16, "D": 14, "E": 14, "F": 36, "G": 26}.items():
        summary.column_dimensions[column].width = width
    set_number_format(summary, [f"B{row}" for row in range(3, end_row + 2)] + [f"D{row}" for row in range(3, end_row + 2)] + [f"E{row}" for row in range(3, end_row + 2)])
    style_table(summary, end_row + 1, 7)

    for record in records:
        ws = wb.create_sheet(record["borrower"][:31])
        if ledger:
            write_receivable_ledger_sheet(ws, record, workbook.name, record["source_row"], source_is_w)
        else:
            ws["A1"] = f"{record['borrower']} 应收账款"
            rows = [
                ("借款人", record["borrower"]),
                (f"表内应收金额/{unit}", record["amount"]),
                ("借款时间", record["date"]),
                (f"已还金额/{unit}", record["repaid"]),
                (f"待还金额/{unit}", record["remaining"]),
                ("备注", record["note"]),
                ("来源", f"{workbook.name}!第{record['source_row']}行"),
            ]
            for row_idx, (label, value) in enumerate(rows, start=3):
                ws.cell(row_idx, 1).value = label
                ws.cell(row_idx, 2).value = value
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 40
            style_table(ws, 9, 2)
            set_number_format(ws, ["B4", "B6", "B7"])

    set_recalculate_on_open(wb)
    wb.save(output)
    return output


def normalize_receivable_ledgers(workbook: Path, template_sheet: str | None = None) -> None:
    wb = load_workbook(workbook, data_only=False)
    records = {record["borrower"]: record for record in receivable_records_from_summary(wb)}
    source_is_w = sheet_has_w_unit(wb["汇总"]) if "汇总" in wb.sheetnames else True
    if not records:
        raise SystemExit("No receivable summary records found.")

    for name, record in records.items():
        if template_sheet and name == template_sheet:
            continue
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name[:31])
        write_receivable_ledger_sheet(ws, record, None, None, source_is_w)

    ordered = [wb["汇总"]] + [wb[name] for name in records if name in wb.sheetnames]
    wb._sheets = ordered + [sheet for sheet in wb.worksheets if sheet not in ordered]
    set_recalculate_on_open(wb)
    wb.save(workbook)


def convert_asset_workbook_to_yuan(wb) -> list[str]:
    changed: list[str] = []
    main = wb["202603"] if "202603" in wb.sheetnames else wb[wb.sheetnames[0]]
    if sheet_has_w_unit(main):
        multiply_range_by_10000(main, range(2, 19), ["C"], changed)
        multiply_range_by_10000(main, range(20, 25), ["C", "E", "F"], changed)
    for sheet_name, rows, columns in (
        ("流动资产", range(3, 11), ["C", "D", "E"]),
        ("投资资产", range(1, 10), ["C"]),
        ("应收账款", range(3, 14), ["B", "G", "L"]),
        ("负债", range(4, 9), ["B", "D", "E"]),
    ):
        if sheet_name in wb.sheetnames and sheet_has_w_unit(wb[sheet_name]):
            multiply_range_by_10000(wb[sheet_name], rows, columns, changed)
    if "应收账款" in wb.sheetnames and sheet_has_w_unit(wb["应收账款"]):
        multiply_range_by_10000(wb["应收账款"], range(17, 20), ["B", "G"], changed)
    return changed


def convert_receivable_workbook_to_yuan(wb) -> list[str]:
    changed: list[str] = []
    if "汇总" in wb.sheetnames and sheet_has_w_unit(wb["汇总"]):
        multiply_range_by_10000(wb["汇总"], range(3, 8), ["B", "D", "E"], changed)
    return changed


def convert_debt_workbook_to_yuan(wb) -> list[str]:
    changed: list[str] = []
    if "汇总" in wb.sheetnames and sheet_has_w_unit(wb["汇总"]):
        multiply_range_by_10000(wb["汇总"], range(3, 8), ["B", "D", "E"], changed)
    for ws in wb.worksheets:
        if ws.title != "汇总" and sheet_has_w_unit(ws):
            multiply_range_by_10000(ws, [3], ["B", "D", "E"], changed)
            multiply_range_by_10000(ws, [6, 8, 9], ["B"], changed)
    return changed


def convert_units_to_yuan(directory: Path) -> list[tuple[Path, int]]:
    results = []
    for path in sorted(directory.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        wb = load_workbook(path, data_only=False)
        if "负债表" in path.name and "资产负债表" not in path.name:
            changed = convert_debt_workbook_to_yuan(wb)
        elif "应收账款" in path.name:
            changed = convert_receivable_workbook_to_yuan(wb)
        else:
            changed = convert_asset_workbook_to_yuan(wb)
        replace_unit_texts(wb)
        set_recalculate_on_open(wb)
        wb.save(path)
        results.append((path, len(changed)))
    return results


def organize_detail_images(directory: Path) -> list[tuple[Path, Path]]:
    moved: list[tuple[Path, Path]] = []
    for path in detail_image_files(directory, recursive=False):
        text = ocr_image_text(path)
        folder = classify_detail_image(path, text)
        if not folder:
            print(f"skip_unclassified: {path.name}")
            continue
        destination = directory / folder / path.name
        new_path = safe_rename(path, destination)
        moved.append((path, new_path))
    return moved


def rename_detail_images(directory: Path, force: bool = False) -> list[tuple[Path, Path]]:
    renamed: list[tuple[Path, Path]] = []
    for path in detail_image_files(directory, recursive=True):
        if path.parent == directory:
            continue
        if not force and "元" in path.stem and not path.name.startswith(("微信图片", "IMG_")):
            continue
        text = ocr_image_text(path)
        new_name = readable_detail_name(path, text)
        if not new_name:
            print(f"skip_no_amount_name: {path.relative_to(directory)}")
            continue
        destination = path.with_name(new_name)
        new_path = safe_rename(path, destination)
        renamed.append((path, new_path))
    return renamed


def verify_workbook(workbook: Path, main_sheet: str | None) -> None:
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")
    print("zip_test", "OK" if workbook_zip_ok(workbook) else "BAD")
    wb = load_workbook(workbook, data_only=False)
    print("sheets:", wb.sheetnames)
    main = first_main_sheet(wb, main_sheet)
    print("main_sheet:", main.title)

    totals: dict[str, float] = {}
    if "流动资产" in wb.sheetnames:
        ws = wb["流动资产"]
        totals["流动资产"] = sum(
            value
            for value in (ws["C3"].value, ws["C4"].value, ws["E5"].value)
            if isinstance(value, (int, float))
        )
    if "投资资产" in wb.sheetnames:
        ws = wb["投资资产"]
        totals["投资资产"] = sum(
            ws.cell(row, 3).value
            for row in range(1, 10)
            if isinstance(ws.cell(row, 3).value, (int, float))
        )
    if "应收账款" in wb.sheetnames:
        ws = wb["应收账款"]
        if ws["A2"].value == "借款人" and ws["E2"].value:
            for row in range(3, ws.max_row + 1):
                if ws.cell(row, 1).value == "合计" and is_number(ws.cell(row, 5).value):
                    totals["应收账款"] = ws.cell(row, 5).value
                    break
        else:
            totals["应收账款"] = sum(
                value
                for value in (ws["B3"].value, ws["G3"].value, ws["L3"].value, ws["B17"].value, ws["G17"].value)
                if isinstance(value, (int, float))
            )
    if "负债" in wb.sheetnames:
        ws = wb["负债"]
        if ws["A2"].value == "负债平台" and ws["F2"].value:
            for row in range(3, ws.max_row + 1):
                if ws.cell(row, 1).value == "合计":
                    if is_number(ws.cell(row, 4).value):
                        totals["负债本金"] = ws.cell(row, 4).value
                    if is_number(ws.cell(row, 5).value):
                        totals["负债利息"] = ws.cell(row, 5).value
                    if is_number(ws.cell(row, 6).value):
                        totals["负债合计"] = ws.cell(row, 6).value
                    break
        else:
            totals["负债本金"] = sum(
                ws.cell(row, 4).value
                for row in range(4, 9)
                if isinstance(ws.cell(row, 4).value, (int, float))
            )
            totals["负债利息"] = sum(
                ws.cell(row, 5).value
                for row in range(4, 9)
                if isinstance(ws.cell(row, 5).value, (int, float))
            )
    print("manual_totals:", {key: round(value, 4) for key, value in totals.items()})


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    split_parser = subparsers.add_parser("split", help="Create/rebuild category sheets from the main sheet")
    split_parser.add_argument("--workbook", required=True, help="Target workbook path")
    split_parser.add_argument("--reference", help="Optional reference workbook for styles/column widths")
    split_parser.add_argument("--main-sheet", help="Main sheet name")

    update_parser = subparsers.add_parser("update-main", help="Update the main asset workbook from sibling receivable/liability summaries")
    update_parser.add_argument("--directory", help="Month folder containing YYYYMM资产表.xlsx, YYYYMM应收账款.xlsx, and YYYYMM负债表.xlsx")
    update_parser.add_argument("--workbook", help="Target asset workbook path")
    update_parser.add_argument("--receivables", help="Receivable workbook path")
    update_parser.add_argument("--liabilities", help="Liability workbook path")
    update_parser.add_argument("--main-sheet", help="Main sheet name")

    archive_parser = subparsers.add_parser("archive", help="Move spreadsheets into YYYY.M/ or YYYY/ folders")
    archive_parser.add_argument("--directory", required=True, help="BalanceSheet directory")

    export_parser = subparsers.add_parser("export-receivables", help="Export receivables to a separate workbook")
    export_parser.add_argument("--workbook", required=True, help="Source asset/liability workbook")
    export_parser.add_argument("--output", help="Output receivable workbook path")
    export_parser.add_argument("--main-sheet", help="Main sheet name")
    export_parser.add_argument("--ledger", action="store_true", help="Create each borrower sheet as a ledger in yuan")

    ledger_parser = subparsers.add_parser("receivable-ledgers", help="Normalize borrower sheets to ledger format")
    ledger_parser.add_argument("--workbook", required=True, help="Receivable workbook path")
    ledger_parser.add_argument("--template-sheet", help="Existing sheet to leave unchanged, such as 游丹")

    units_parser = subparsers.add_parser("convert-units", help="Convert W/万元 amount fields to yuan in a directory")
    units_parser.add_argument("--directory", required=True, help="Directory containing workbooks")

    details_parser = subparsers.add_parser("organize-details", help="Move detail screenshots into platform folders")
    details_parser.add_argument("--directory", required=True, help="Detail screenshot directory")

    rename_parser = subparsers.add_parser("rename-detail-images", help="Rename detail screenshots with visible amount names")
    rename_parser.add_argument("--directory", required=True, help="Detail screenshot directory")
    rename_parser.add_argument("--force", action="store_true", help="Rename files even if the current name already contains 元")

    verify_parser = subparsers.add_parser("verify", help="Validate workbook readability and print key totals")
    verify_parser.add_argument("--workbook", required=True, help="Target workbook path")
    verify_parser.add_argument("--main-sheet", help="Main sheet name")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    if args.command == "split":
        workbook = normalize_path(args.workbook)
        reference = normalize_path(args.reference) if getattr(args, "reference", None) else None
        split_categories(workbook, reference, args.main_sheet)
        print(f"updated: {workbook}")
        verify_workbook(workbook, args.main_sheet)
    elif args.command == "update-main":
        directory = normalize_path(args.directory) if args.directory else None
        workbook = normalize_path(args.workbook) if args.workbook else infer_asset_workbook(directory) if directory else None
        receivables = normalize_path(args.receivables) if args.receivables else infer_month_workbook(directory, "应收账款") if directory else None
        liabilities = normalize_path(args.liabilities) if args.liabilities else infer_month_workbook(directory, "负债表") if directory else None
        if workbook is None or receivables is None or liabilities is None:
            raise SystemExit("Pass --directory, or pass --workbook plus --receivables plus --liabilities.")
        update_main_from_summaries(workbook, receivables, liabilities, args.main_sheet)
        verify_workbook(workbook, args.main_sheet)
    elif args.command == "archive":
        directory = normalize_path(args.directory)
        moved = archive_by_date(directory)
        for source, destination in moved:
            print(f"moved: {source} -> {destination}")
    elif args.command == "export-receivables":
        workbook = normalize_path(args.workbook)
        output = normalize_path(args.output) if args.output else None
        exported = export_receivables(workbook, output, args.main_sheet, args.ledger)
        print(f"exported: {exported}")
        print("zip_test", "OK" if workbook_zip_ok(exported) else "BAD")
    elif args.command == "receivable-ledgers":
        workbook = normalize_path(args.workbook)
        normalize_receivable_ledgers(workbook, args.template_sheet)
        print(f"updated: {workbook}")
        print("zip_test", "OK" if workbook_zip_ok(workbook) else "BAD")
    elif args.command == "convert-units":
        directory = normalize_path(args.directory)
        for path, count in convert_units_to_yuan(directory):
            print(f"converted: {path.name} cells={count} zip={'OK' if workbook_zip_ok(path) else 'BAD'}")
    elif args.command == "organize-details":
        directory = normalize_path(args.directory)
        for source, destination in organize_detail_images(directory):
            print(f"moved: {source.relative_to(directory)} -> {destination.relative_to(directory)}")
    elif args.command == "rename-detail-images":
        directory = normalize_path(args.directory)
        for source, destination in rename_detail_images(directory, args.force):
            print(f"renamed: {source.relative_to(directory)} -> {destination.relative_to(directory)}")
    elif args.command == "verify":
        workbook = normalize_path(args.workbook)
        verify_workbook(workbook, args.main_sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
